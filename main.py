from __future__ import annotations

import os
import uuid
import shutil
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Optional, Type

from fastapi import Depends, FastAPI, File, HTTPException, Query, UploadFile, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm

from pydantic import BaseModel, ConfigDict, create_model
from pydantic_settings import BaseSettings

from sqlalchemy import create_engine, Column, String, Integer, Boolean, Float, Text, DateTime, Date, Numeric, select, func, or_
from sqlalchemy.orm import declarative_base, sessionmaker, Session, relationship
from sqlalchemy.sql import text

from passlib.context import CryptContext
from jose import jwt, JWTError

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import cm

# ============================================================================
# CONFIG
# ============================================================================
class Settings(BaseSettings):
    DATABASE_URL: str = "sqlite:///./shopping_security_cmms.db"
    SECRET_KEY: str = "shopping-security-cmms-secret-key-change-in-production"
    ALGORITHM: str = "HS256"
    ACCESS_TOKEN_EXPIRE: int = 480
    UPLOAD_DIR: str = "uploads"
    APP_NAME: str = "Shopping Security CMMS"
    CORS_ORIGINS: str = "*"

    class Config:
        env_file = ".env"
        extra = "ignore"

settings = Settings()
os.makedirs(settings.UPLOAD_DIR, exist_ok=True)

# ============================================================================
# DATABASE
# ============================================================================
connect_args = {"check_same_thread": False} if settings.DATABASE_URL.startswith("sqlite") else {}
engine = create_engine(settings.DATABASE_URL, connect_args=connect_args, pool_pre_ping=True)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# ============================================================================
# BASE MIXIN
# ============================================================================
class BaseModelMixin:
    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)
    created_by = Column(String(36), nullable=True)
    updated_by = Column(String(36), nullable=True)

    def __repr__(self):
        name = getattr(self, "nome", None) or getattr(self, "codigo", None) or self.__class__.__name__
        return f"<{self.__class__.__name__} id={self.id} nome={name}>"

# ============================================================================
# MODELS
# ============================================================================
class Usuario(Base, BaseModelMixin):
    __tablename__ = "usuarios"
    nome = Column(String(255), nullable=False)
    email = Column(String(255), unique=True, nullable=False, index=True)
    senha_hash = Column(String(255), nullable=False)
    perfil = Column(String(50), default="operador", nullable=False)
    ativo = Column(Boolean, default=True, nullable=False)
    telefone = Column(String(20))
    cargo = Column(String(100))

class Auditoria(Base, BaseModelMixin):
    __tablename__ = "auditoria"
    usuario_id = Column(String(36), nullable=True)
    entidade = Column(String(100), nullable=False)
    entidade_id = Column(String(36), nullable=True)
    acao = Column(String(50), nullable=False)
    detalhes = Column(Text)
    valores_anteriores = Column(Text)
    valores_novos = Column(Text)

class Extintor(Base, BaseModelMixin):
    __tablename__ = "extintores"
    numero_extintor = Column(String(50), nullable=False)
    patrimonio = Column(String(50))
    classe_incendio = Column(String(10))
    tipo = Column(String(50), nullable=False)
    capacidade = Column(String(20))
    fabricante = Column(String(100))
    numero_serie = Column(String(100))
    data_fabricacao = Column(Date)
    data_teste_hidrostatico = Column(Date)
    data_validade = Column(Date, index=True)
    localizacao = Column(String(255))
    pavimento = Column(String(50))
    setor = Column(String(50))
    foto_url = Column(String(255))
    situacao = Column(String(30), default="conforme")

class Mangueira(Base, BaseModelMixin):
    __tablename__ = "mangueiras"
    numero_mangueira = Column(String(50), nullable=False)
    tipo = Column(String(20))
    comprimento = Column(String(20))
    diametro = Column(String(20))
    fabricante = Column(String(100))
    data_fabricacao = Column(Date)
    data_ultimo_ensaio = Column(Date)
    data_validade = Column(Date, index=True)
    abrigo_hidrante = Column(String(50))
    situacao = Column(String(30), default="conforme")
    foto_url = Column(String(255))

class Abrigo(Base, BaseModelMixin):
    __tablename__ = "abrigos"
    codigo_abrigo = Column(String(50), nullable=False)
    localizacao = Column(String(255))
    pavimento = Column(String(50))
    setor = Column(String(50))
    foto_url = Column(String(255))
    tem_mangueira = Column(String(20), default="conforme")
    tem_chave_storz = Column(String(20), default="conforme")
    tem_esguicho = Column(String(20), default="conforme")
    tem_tampao = Column(String(20), default="conforme")
    tem_reducao = Column(String(20), default="conforme")
    tem_adaptadores = Column(String(20), default="conforme")
    tem_registro = Column(String(20), default="conforme")
    tem_sinalizacao = Column(String(20), default="conforme")
    outros_equipamentos = Column(Text)

class EquipamentoCombate(Base, BaseModelMixin):
    __tablename__ = "equipamentos_combate"
    tipo = Column(String(100), nullable=False)
    patrimonio = Column(String(50))
    quantidade = Column(Integer, default=1)
    localizacao = Column(String(255))
    estado_conservacao = Column(String(30), default="bom")
    foto_url = Column(String(255))
    data_inspecao = Column(Date)

class KitCrise(Base, BaseModelMixin):
    __tablename__ = "kits_crise"
    item = Column(String(255), nullable=False)
    quantidade = Column(Integer, default=0)
    localizacao = Column(String(255))
    responsavel = Column(String(255))
    situacao = Column(String(30), default="conforme")
    validade = Column(Date)

class VGA(Base, BaseModelMixin):
    __tablename__ = "vgas"
    numero_vga = Column(String(50), nullable=False)
    localizacao = Column(String(255))
    pavimento = Column(String(50))
    vazao = Column(String(50))
    pressao = Column(String(50))
    data_ultima_inspecao = Column(Date)
    situacao = Column(String(30), default="conforme")
    fotos = Column(Text)
    historico_manutencao = Column(Text)

class Loja(Base, BaseModelMixin):
    __tablename__ = "lojas"
    nome_loja = Column(String(255), nullable=False)
    numero = Column(String(20))
    segmento = Column(String(100))
    responsavel = Column(String(255))
    telefone = Column(String(20))
    email = Column(String(255))
    sistema_incendio = Column(Text)
    sistema_deteccao = Column(Text)
    sprinklers = Column(Text)
    hidrantes = Column(Text)
    extintores = Column(Text)
    cftv = Column(Text)
    central_alarme = Column(Text)
    problemas_encontrados = Column(Text)
    fotos = Column(Text)
    data_vistoria = Column(Date)
    responsavel_vistoria = Column(String(255))
    grau_risco = Column(String(20), default="Risco 01")

class Notificacao(Base, BaseModelMixin):
    __tablename__ = "notificacoes"
    numero_notificacao = Column(String(50), nullable=False)
    loja = Column(String(255))
    irregularidade = Column(Text)
    fotos = Column(Text)
    data_emissao = Column(Date)
    prazo_regularizacao = Column(Date)
    status = Column(String(30), default="aberta")
    historico_interacoes = Column(Text)
    arquivos_anexos = Column(Text)

class Sprinkler(Base, BaseModelMixin):
    __tablename__ = "sprinklers"
    codigo = Column(String(50), nullable=False)
    bombas = Column(Text)
    valvulas = Column(Text)
    fluxostatos = Column(Text)
    chaves_fluxo = Column(Text)
    pressostatos = Column(Text)
    rede_hidraulica = Column(Text)
    setores = Column(Text)
    data_inspecao = Column(Date)
    testes_realizados = Column(Text)
    nao_conformidades = Column(Text)
    fotos = Column(Text)
    historico_manutencao = Column(Text)

class SDAI(Base, BaseModelMixin):
    __tablename__ = "sdai"
    codigo = Column(String(50), nullable=False)
    central_alarme = Column(String(255))
    acionadores_manuais = Column(Text)
    detectores = Column(Text)
    sirenes = Column(Text)
    modulos = Column(Text)
    fontes = Column(Text)
    baterias = Column(Text)
    loop = Column(String(50))
    enderecamento = Column(Text)
    testes = Column(Text)
    falhas = Column(Text)
    manutencoes = Column(Text)
    fotos = Column(Text)
    historico = Column(Text)

class CFTV(Base, BaseModelMixin):
    __tablename__ = "cftv"
    codigo = Column(String(50), nullable=False)
    tipo = Column(String(50))
    marca = Column(String(100))
    modelo = Column(String(100))
    ip = Column(String(50))
    localizacao = Column(String(255))
    estado = Column(String(30), default="ativo")
    garantia = Column(Date)
    data_manutencao = Column(Date)
    fotos = Column(Text)

class ControleAcesso(Base, BaseModelMixin):
    __tablename__ = "controle_acesso"
    codigo = Column(String(50), nullable=False)
    tipo = Column(String(50))
    marca = Column(String(100))
    modelo = Column(String(100))
    localizacao = Column(String(255))
    manutencoes = Column(Text)
    falhas = Column(Text)
    historico = Column(Text)

class Gerador(Base, BaseModelMixin):
    __tablename__ = "geradores"
    codigo = Column(String(50), nullable=False)
    horimetro = Column(String(50))
    testes_semanais = Column(Date)
    testes_mensais = Column(Date)
    nivel_combustivel = Column(String(50))
    manutencoes = Column(Text)
    baterias = Column(Text)
    alarmes = Column(Text)

class BombaIncendio(Base, BaseModelMixin):
    __tablename__ = "bombas_incendio"
    codigo = Column(String(50), nullable=False)
    tipo = Column(String(50))
    bomba_principal = Column(Text)
    jockey = Column(Text)
    reserva = Column(Text)
    testes = Column(Text)
    pressao = Column(String(50))
    vazao = Column(String(50))
    manutencoes = Column(Text)
    falhas = Column(Text)

class IluminacaoEmergencia(Base, BaseModelMixin):
    __tablename__ = "iluminacao_emergencia"
    codigo = Column(String(50), nullable=False)
    tipo = Column(String(50))
    localizacao = Column(String(255))
    data_teste = Column(Date)
    autonomia = Column(String(50))
    situacao = Column(String(30), default="conforme")

class SinalizacaoEmergencia(Base, BaseModelMixin):
    __tablename__ = "sinalizacao_emergencia"
    codigo = Column(String(50), nullable=False)
    tipo = Column(String(50))
    estado_conservacao = Column(String(30), default="bom")
    fotos = Column(Text)
    pendencias = Column(Text)

class PortaCortafogo(Base, BaseModelMixin):
    __tablename__ = "portas_cortafogo"
    codigo = Column(String(50), nullable=False)
    localizacao = Column(String(255))
    barra_antipanico = Column(String(20), default="conforme")
    fechadura = Column(String(20), default="conforme")
    molas = Column(String(20), default="conforme")
    situacao = Column(String(30), default="conforme")
    data_inspecao = Column(Date)

class Brigada(Base, BaseModelMixin):
    __tablename__ = "brigada"
    nome = Column(String(255), nullable=False)
    cargo = Column(String(100))
    escala = Column(Text)
    treinamentos = Column(Text)
    reciclagens = Column(Text)
    certificados = Column(Text)
    data_validade = Column(Date, index=True)

class Documento(Base, BaseModelMixin):
    __tablename__ = "documentos"
    titulo = Column(String(255), nullable=False)
    tipo = Column(String(50))
    data_emissao = Column(Date)
    data_validade = Column(Date, index=True)
    arquivo_url = Column(String(255))
    observacoes = Column(Text)

class Ocorrencia(Base, BaseModelMixin):
    __tablename__ = "ocorrencias"
    tipo = Column(String(50))
    data = Column(Date)
    hora = Column(String(10))
    local = Column(String(255))
    descricao = Column(Text)
    fotos = Column(Text)
    responsavel = Column(String(255))
    providencias = Column(Text)

class Cronograma(Base, BaseModelMixin):
    __tablename__ = "cronograma"
    titulo = Column(String(255), nullable=False)
    periodicidade = Column(String(30))
    data_programada = Column(Date)
    data_realizacao = Column(Date)
    responsavel = Column(String(255))
    observacoes = Column(Text)

class Alerta(Base, BaseModelMixin):
    __tablename__ = "alertas"
    titulo = Column(String(255), nullable=False)
    entidade = Column(String(100))
    entidade_id = Column(String(36))
    dias_restantes = Column(Integer)
    nivel = Column(String(20), default="media")
    status = Column(String(20), default="aberto")

# ============================================================================
# AUTH UTILS
# ============================================================================
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/auth/login")

def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=settings.ACCESS_TOKEN_EXPIRE)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.SECRET_KEY, algorithm=settings.ALGORITHM)

def decode_token(token: str) -> dict:
    try:
        return jwt.decode(token, settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
    except JWTError:
        raise HTTPException(status_code=401, detail="Token inválido ou expirado")

def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    payload = decode_token(token)
    user_id = payload.get("sub")
    if not user_id:
        raise HTTPException(status_code=401, detail="Token inválido")
    user = db.query(Usuario).filter(Usuario.id == user_id).first()
    if not user or not user.ativo:
        raise HTTPException(status_code=401, detail="Usuário não encontrado ou inativo")
    return user

# ============================================================================
# AUDITORIA SERVICE
# ============================================================================
def registrar_acao(db, usuario_id, entidade, entidade_id, acao, detalhes=None, valores_anteriores=None, valores_novos=None):
    reg = Auditoria(
        usuario_id=usuario_id,
        entidade=entidade,
        entidade_id=entidade_id,
        acao=acao,
        detalhes=detalhes,
        valores_anteriores=valores_anteriores,
        valores_novos=valores_novos,
    )
    db.add(reg)
    db.commit()

# ============================================================================
# SCHEMAS
# ============================================================================
class SchemaBase(BaseModel):
    model_config = ConfigDict(from_attributes=True)

# ============================================================================
# DASHBOARD SERVICE
# ============================================================================
def verificar_vencimentos(db):
    hoje = date.today()
    modelos_data = [
        (Extintor, "data_validade"),
        (Mangueira, "data_validade"),
        (Brigada, "data_validade"),
        (Documento, "data_validade"),
        (KitCrise, "validade"),
    ]
    vencidos = []
    proximos_30 = []
    proximos_60 = []
    proximos_90 = []
    for model, campo in modelos_data:
        col = getattr(model, campo, None)
        if not col:
            continue
        itens = db.query(model).filter(col.isnot(None)).all()
        for item in itens:
            data_val = getattr(item, campo)
            if not data_val:
                continue
            dias = (data_val - hoje).days
            info = {
                "entidade": model.__tablename__,
                "id": item.id,
                "nome": getattr(item, "nome", None) or getattr(item, "numero_extintor", None) or getattr(item, "codigo", None) or str(item.id),
                "data": str(data_val),
                "dias_restantes": dias,
            }
            if dias < 0:
                vencidos.append(info)
            elif dias <= 30:
                proximos_30.append(info)
            elif dias <= 60:
                proximos_60.append(info)
            elif dias <= 90:
                proximos_90.append(info)
    return {
        "vencidos": vencidos,
        "proximos_30_dias": proximos_30,
        "proximos_60_dias": proximos_60,
        "proximos_90_dias": proximos_90,
    }

def get_kpis(db):
    hoje = date.today()
    total_extintores = db.query(func.count(Extintor.id)).scalar() or 0
    total_mangueiras = db.query(func.count(Mangueira.id)).scalar() or 0
    total_equipamentos = total_extintores + total_mangueiras
    vencidos = db.query(func.count(Extintor.id)).filter(Extintor.data_validade < hoje).scalar() or 0
    vencidos += db.query(func.count(Mangueira.id)).filter(Mangueira.data_validade < hoje).scalar() or 0
    proximos = db.query(func.count(Extintor.id)).filter(
        Extintor.data_validade >= hoje, Extintor.data_validade <= hoje + timedelta(days=30)
    ).scalar() or 0
    notif_abertas = db.query(func.count(Notificacao.id)).filter(Notificacao.status.in_(["aberta", "em andamento"])).scalar() or 0
    ocorrencias_abertas = db.query(func.count(Ocorrencia.id)).filter(Ocorrencia.providencias == None).scalar() or 0
    return {
        "total_equipamentos": total_equipamentos,
        "vencidos": vencidos,
        "proximos_30_dias": proximos,
        "notificacoes_abertas": notif_abertas,
        "ocorrencias_pendentes": ocorrencias_abertas,
    }

# ============================================================================
# RELATORIO SERVICE
# ============================================================================
def gerar_relatorio_pdf(tipo, dados):
    import io
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    elementos = []
    elementos.append(Paragraph("Shopping Security CMMS", styles["Title"]))
    elementos.append(Paragraph(f"Relatório de {tipo.capitalize()}", styles["Heading2"]))
    elementos.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
    elementos.append(Spacer(1, 0.5*cm))
    if dados:
        colunas = list(dados[0].keys())
        tabela = [colunas] + [[str(row.get(c, "")) for c in colunas] for row in dados]
        t = Table(tabela, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a237e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f5f5f5")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elementos.append(t)
    doc.build(elementos)
    return buffer.getvalue()

def gerar_relatorio_excel(tipo, dados):
    import io
    buffer = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tipo[:31]
    ws["A1"] = "Shopping Security CMMS"
    ws["A2"] = f"Relatório de {tipo.capitalize()}"
    ws["A3"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"].font = Font(bold=True, size=12)
    if dados:
        colunas = list(dados[0].keys())
        for ci, col in enumerate(colunas, 1):
            cell = ws.cell(row=5, column=ci, value=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
        for ri, row in enumerate(dados, 6):
            for ci, col in enumerate(colunas, 1):
                ws.cell(row=ri, column=ci, value=str(row.get(col, "")))
    wb.save(buffer)
    return buffer.getvalue()

# ============================================================================
# GENERIC CRUD ROUTER FACTORY
# ============================================================================
def build_crud_router(prefix, model, create_schema=None, update_schema=None):
    router = FastAPI()

    @router.get("/")
    def listar(skip: int = 0, limit: int = 100, search: Optional[str] = None, db: Session = Depends(get_db), user=Depends(get_current_user)):
        q = db.query(model)
        if search:
            string_cols = [c.name for c in model.__table__.columns if isinstance(c.type, (String, Text)) and c.name not in ("id", "senha_hash")]
            if string_cols:
                q = q.filter(or_(*[getattr(model, c).ilike(f"%{search}%") for c in string_cols]))
        return q.order_by(model.created_at.desc()).offset(skip).limit(limit).all()

    @router.get("/{item_id}")
    def obter(item_id: str, db: Session = Depends(get_db), user=Depends(get_current_user)):
        item = db.query(model).filter(model.id == item_id).first()
        if not item:
            raise HTTPException(status_code=404, detail="Registro não encontrado")
        return item

    @router.post("/")
    def criar(payload: dict, db: Session = Depends(get_db), user=Depends(get_current_user)):
        item = model(**payload, created_by=user.id, updated_by=user.id)
        db.add(item)
        db.commit()
        db.refresh(item)
        registrar_acao(db, user.id, prefix, item.id, "create", valores_novos=str(payload))
        return item

    @router.put("/{item_id}")
    def atualizar(item_id: str, payload: dict, db: Session = Depends(get_db), user=Depends(get_current_user)):
        item = db.query(model).filter(model.id == item_id).first()
        if not item:
            raise HTTPException(status_code=404, detail="Registro não encontrado")
        anterior = {c.name: getattr(item, c.name) for c in model.__table__.columns}
        for k, v in payload.items():
            setattr(item, k, v)
        item.updated_by = user.id
        db.commit()
        db.refresh(item)
        registrar_acao(db, user.id, prefix, item.id, "update", valores_anteriores=str(anterior), valores_novos=str(payload))
        return item

    @router.delete("/{item_id}")
    def excluir(item_id: str, db: Session = Depends(get_db), user=Depends(get_current_user)):
        item = db.query(model).filter(model.id == item_id).first()
        if not item:
            raise HTTPException(status_code=404, detail="Registro não encontrado")
        db.delete(item)
        db.commit()
        registrar_acao(db, user.id, prefix, item.id, "delete")
        return {"detail": "Registro excluído com sucesso"}

    return router

# ============================================================================
# APP
# ============================================================================
app = FastAPI(
    title=settings.APP_NAME,
    description="CMMS completo para Segurança Patrimonial e Brigada de Incêndio de Shopping Center",
    version="1.0.0",
    docs_url="/docs",
    redoc_url="/redoc",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/uploads", StaticFiles(directory=settings.UPLOAD_DIR), name="uploads")

@app.on_event("startup")
def on_startup():
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        if not db.query(Usuario).filter(Usuario.email == "admin@shopping.com").first():
            admin = Usuario(
                nome="Administrador",
                email="admin@shopping.com",
                senha_hash=hash_password("admin123"),
                perfil="admin",
                cargo="Gestor de Segurança",
            )
            db.add(admin)
            db.commit()
    finally:
        db.close()

@app.get("/health")
async def health():
    return {"status": "ok", "app": settings.APP_NAME, "time": datetime.utcnow().isoformat()}

@app.get("/")
async def root():
    return {"app": settings.APP_NAME, "docs": "/docs", "version": "1.0.0"}

# ============================================================================
# AUTH ROUTES
# ============================================================================
@app.post("/auth/login")
async def login(form: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(Usuario).filter(Usuario.email == form.username).first()
    if not user or not verify_password(form.password, user.senha_hash):
        raise HTTPException(status_code=401, detail="Credenciais inválidas")
    if not user.ativo:
        raise HTTPException(status_code=403, detail="Usuário inativo")
    token = create_access_token({"sub": user.id, "perfil": user.perfil})
    return {"access_token": token, "token_type": "bearer", "user_id": user.id, "nome": user.nome, "perfil": user.perfil}

@app.post("/auth/register")
async def register(nome: str, email: str, senha: str, perfil: str = "operador", db: Session = Depends(get_db)):
    if db.query(Usuario).filter(Usuario.email == email).first():
        raise HTTPException(status_code=400, detail="Email já cadastrado")
    user = Usuario(nome=nome, email=email, senha_hash=hash_password(senha), perfil=perfil)
    db.add(user)
    db.commit()
    db.refresh(user)
    return user

@app.get("/auth/me")
async def me(user=Depends(get_current_user)):
    return user

# ============================================================================
# DASHBOARD ROUTES
# ============================================================================
@app.get("/dashboard/kpis")
async def dashboard_kpis(db: Session = Depends(get_db), user=Depends(get_current_user)):
    return get_kpis(db)

@app.get("/dashboard/vencimentos")
async def dashboard_vencimentos(db: Session = Depends(get_db), user=Depends(get_current_user)):
    return verificar_vencimentos(db)

@app.get("/dashboard/lojas-risco")
async def dashboard_lojas_risco(db: Session = Depends(get_db), user=Depends(get_current_user)):
    lojas = db.query(Loja).all()
    distribuicao = {"Risco 01": 0, "Risco 02": 0, "Risco 03": 0, "Risco 04": 0}
    for loja in lojas:
        distribuicao[loja.grau_risco] = distribuicao.get(loja.grau_risco, 0) + 1
    return {"distribuicao": distribuicao, "lojas": lojas}

# ============================================================================
# RELATORIO ROUTES
# ============================================================================
@app.post("/relatorios/pdf")
async def relatorio_pdf(tipo: str = Query(...), db: Session = Depends(get_db), user=Depends(get_current_user)):
    modelos = {
        "extintores": Extintor,
        "mangueiras": Mangueira,
        "abrigos": Abrigo,
        "lojas": Loja,
        "notificacoes": Notificacao,
        "ocorrencias": Ocorrencia,
        "brigada": Brigada,
        "documentos": Documento,
    }
    model = modelos.get(tipo)
    if not model:
        raise HTTPException(status_code=400, detail="Tipo inválido")
    rows = db.query(model).all()
    dados = [{c.name: str(getattr(r, c.name)) for c in model.__table__.columns} for r in rows]
    pdf = gerar_relatorio_pdf(tipo, dados)
    return Response(content=pdf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=relatorio_{tipo}.pdf"})

@app.post("/relatorios/excel")
async def relatorio_excel(tipo: str = Query(...), db: Session = Depends(get_db), user=Depends(get_current_user)):
    modelos = {
        "extintores": Extintor,
        "mangueiras": Mangueira,
        "abrigos": Abrigo,
        "lojas": Loja,
        "notificacoes": Notificacao,
        "ocorrencias": Ocorrencia,
        "brigada": Brigada,
        "documentos": Documento,
    }
    model = modelos.get(tipo)
    if not model:
        raise HTTPException(status_code=400, detail="Tipo inválido")
    rows = db.query(model).all()
    dados = [{c.name: str(getattr(r, c.name)) for c in model.__table__.columns} for r in rows]
    xlsx = gerar_relatorio_excel(tipo, dados)
    return Response(content=xlsx, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename=relatorio_{tipo}.xlsx"})

# ============================================================================
# MÓDULOS CRUD
# ============================================================================
MODULOS = [
    ("extintores", Extintor),
    ("mangueiras", Mangueira),
    ("abrigos", Abrigo),
    ("equipamentos-combate", EquipamentoCombate),
    ("kits-crise", KitCrise),
    ("vgas", VGA),
    ("lojas", Loja),
    ("notificacoes", Notificacao),
    ("sprinklers", Sprinkler),
    ("sdai", SDAI),
    ("cftv", CFTV),
    ("controle-acesso", ControleAcesso),
    ("geradores", Gerador),
    ("bombas-incendio", BombaIncendio),
    ("iluminacao-emergencia", IluminacaoEmergencia),
    ("sinalizacao-emergencia", SinalizacaoEmergencia),
    ("portas-cortafogo", PortaCortafogo),
    ("brigada", Brigada),
    ("documentos", Documento),
    ("ocorrencias", Ocorrencia),
    ("cronograma", Cronograma),
    ("alertas", Alerta),
]

for prefixo, modelo in MODULOS:
    router = build_crud_router(prefixo, modelo)
    app.mount(f"/{prefixo}", router)

# ============================================================================
# UPLOAD
# ============================================================================
@app.post("/upload")
async def upload_file(file: UploadFile = File(...), user=Depends(get_current_user)):
    ext = (file.filename or "").rsplit(".", 1)[-1].lower()
    if ext not in {"jpg", "jpeg", "png", "webp", "pdf"}:
        raise HTTPException(status_code=400, detail="Formato não permitido")
    nome = f"{uuid.uuid4().hex}.{ext}"
    path = os.path.join(settings.UPLOAD_DIR, nome)
    with open(path, "wb") as f:
        shutil.copyfileobj(file.file, f)
    return {"url": f"/uploads/{nome}", "filename": nome}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
